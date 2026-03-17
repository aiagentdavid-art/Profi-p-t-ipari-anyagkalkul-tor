import os
import google.generativeai as genai
from dotenv import load_dotenv

load_dotenv()

def analyze_file(file_path, file_type):
    """
    Elemzi a feltöltött fájlt (PDF, Kép, Docx, Xlsx) a Gemini AI segítségével.
    """
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        return {"error": "Hiányzó API kulcs."}

    genai.configure(api_key=api_key)
    
    try:
        # Szöveges tartalom kinyerése, ha Excel
        excel_text_content = ""
        uploaded_file = None
        
        if str(file_path).lower().endswith('.xlsx'):
            import openpyxl
            try:
                print("Excel fájl olvasása openpyxl segítségével...")
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    excel_text_content += f"\\n--- Munkalap: {sheet_name} ---\\n"
                    for row in sheet.iter_rows(values_only=True):
                        # Csak a nem üres sorokat adjuk hozzá
                        if any(cell is not None for cell in row):
                            row_str = " | ".join([str(c) if c is not None else "" for c in row])
                            excel_text_content += row_str + "\\n"
            except Exception as e:
                print(f"Hiba az Excel olvasásakor: {e}")
                # Fallback, ha nem ment
                excel_text_content = ""
        
        # Ha nem Excel, vagy ha fallback kell, feltöltjük a fájlt
        if not excel_text_content:
            print(f"Fájl feltöltése a Geminihez: {file_path}")
            uploaded_file = genai.upload_file(path=file_path)
        
        # Modell inicializálása (3.1 Flash javasolt a vegyes adatokhoz)
        model = genai.GenerativeModel(model_name="gemini-3.1-flash-lite-preview")
        
        prompt = f"""
        Elemezd a csatolt építőipari dokumentumot (alaprajz, ajánlat, táblázat).
        MINDEN oldalt és munkalapot (sheet) vizsgálj meg, mert a lényeges adatok lehet, hogy nem az elején vannak!
        {"Íme az Excel fájl kinyert adatai:" + excel_text_content if excel_text_content else ""}
        
        Keresd ki a számításokhoz szükséges adatokat (a megnevezések változhatnak, pl. 'gipszkarton fal', 'hidegburkolás', 'festés', stb.):
        - Felületek, mennyiségek (m2, db, fm)
        
        FONTOS SZABÁLYOK:
        1. NE nyerj ki végösszegeket, árakat, forint értékeket, sem "munkadíj" sem "anyagár" összesítőt! Csak a FIZIKAI anyagmennyiségekre koncentrálj!
        2. Ha a szövegben szerepel a rétegszám (pl. "2 rétegű"), azt MINDIG írd bele a kulcs nevébe!
        3. Ha a szövegben szerepel a csempe, burkolat mérete (pl. "30*60", "30x60"), azt MINDIG írd bele a kulcs nevébe!

        Válaszolj kizárólag egy JSON formátumban, ahol a kulcs a RÉSZLETES megnevezés, az érték pedig csak a szám (mennyiség).
        Példa: {{"Nappali felület (m2)": 25.5, "Gipszkarton válaszfal 2 réteg (m2)": 15, "Hidegburkolás 30x60 lap (db)": 12}}
        Ha nem találsz adatot, küldj üres objektumot.
        """
        
        # Ha van excel string, csak a prompt megy, ha nincs, a prompt + file
        contents = [prompt]
        if uploaded_file:
            contents.append(uploaded_file)
            
        response = model.generate_content(contents)
        
        # Tisztítás
        text = response.text
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            text = text.split("```")[1].split("```")[0]
            
        import json
        return json.loads(text.strip())
        
    except Exception as e:
        return {"error": f"Fájl elemzési hiba: {str(e)}"}
    finally:
        # Ideiglenes fájl törlése (opcionális, a hívó fél kezeli)
        pass
